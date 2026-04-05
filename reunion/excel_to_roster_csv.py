"""
excel_to_roster_csv.py - 出席表Excelを名簿CSVに変換するスクリプト

使い方:
  python excel_to_roster_csv.py <Excelファイルパス> [出力CSVパス]

例:
  python excel_to_roster_csv.py "出席表.xlsx"
  python excel_to_roster_csv.py "出席表.xlsx" data/roster.csv

出力CSV列:
  氏名, メールアドレス, クラス, 出席番号, 役割, 幹事メモ

クラス番号ルール:
  - 2桁の数字: 十の位=学年, 一の位=組 (例: 31=3年1組, 39=3年9組)
  - 学年主任: クラス列は空、役割=学年主任
  - 担任・副担任: 担当クラスの番号、役割=教師
  - 生徒: 所属クラスの番号、役割=生徒
"""
import sys
import csv
import re
import openpyxl
from pathlib import Path


# ──────────────────────────────────────────
# Excelレイアウト定数
# ──────────────────────────────────────────
# 1行目: ヘッダー（クラス名など）
# 2行目〜: 生徒データ（1列目に「生徒」ラベル）
# 43行目〜: 担任・副担任データ（1列目に「担任・副担任」ラベル）
# 各クラスの名前列（1-indexed）: 31組=2, 32組=8, ... (6列ごと)
STUDENT_START_ROW = 2      # 生徒データ開始行
TEACHER_START_ROW = 43     # 担任・副担任開始行
FIRST_NAME_COL = 2         # 最初のクラス(31組)の名前列インデックス(1-indexed)
CLASS_COL_STEP = 6         # クラス間の列数

# クラス番号リスト (2桁)
CLASS_NUMBERS = [31, 32, 33, 34, 35, 36, 37, 38, 39]


def get_class_col(class_number: int) -> int:
    """クラス番号から名前列インデックス(1-indexed)を返す"""
    idx = CLASS_NUMBERS.index(class_number)
    return FIRST_NAME_COL + idx * CLASS_COL_STEP


def is_valid_name(value) -> bool:
    """有効な名前かどうか判定（None・空文字・集計用の値を除外）"""
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    # 数字のみ・記号のみは名前として無効
    if re.fullmatch(r'[\d\s○×△]+', s):
        return False
    # 集計ラベルを除外
    skip_words = {"計", "合計", "出席情報", "参加", "不参加", "未定",
                  "生徒", "担任・副担任", "仮出欠", "本出欠", "当日",
                  "reメール", "連絡先"}
    if s.lower() in skip_words:
        return False
    return True


def parse_excel(filepath: str) -> list[dict]:
    """
    Excelを解析して参加者リストを返す。

    Returns:
        list[dict]: 各参加者の辞書
            - name: str
            - email: str (空)
            - class_name: str (2桁数字 or 空)
            - student_number: str
            - role: str (生徒/教師/学年主任)
            - teacher_memo: str
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Sheet1 を使用（なければ最初のシート）
    sheet_name = "Sheet1" if "Sheet1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    print(f"シート「{sheet_name}」を解析中... ({ws.max_row}行 x {ws.max_column}列)")

    records = []
    seen_names = set()  # 重複名チェック用

    def add_record(name, class_name, student_number, role, memo=""):
        name = str(name).strip()
        if not is_valid_name(name):
            return
        key = (name, class_name, role)
        if key in seen_names:
            return
        seen_names.add(key)
        records.append({
            "name": name,
            "email": "",
            "class_name": str(class_name),
            "student_number": str(student_number) if student_number else "",
            "role": role,
            "teacher_memo": memo,
        })

    # ──────────────────────────────────────────
    # 生徒の抽出
    # ──────────────────────────────────────────
    for class_num in CLASS_NUMBERS:
        col = get_class_col(class_num)
        student_count = 0

        for row_idx in range(STUDENT_START_ROW, TEACHER_START_ROW):
            cell_val = ws.cell(row=row_idx, column=col).value
            if is_valid_name(cell_val):
                student_count += 1
                add_record(
                    name=cell_val,
                    class_name=str(class_num),
                    student_number=student_count,
                    role="生徒",
                )

        print(f"  {class_num}組: 生徒 {student_count} 名")

    # ──────────────────────────────────────────
    # 担任・副担任の抽出
    # ──────────────────────────────────────────
    # 行43以降、各クラス列の名前を教師として取込む
    teacher_rows_found = 0
    for row_idx in range(TEACHER_START_ROW, ws.max_row + 1):
        # 行の最初のセルを確認（ラベル行かデータ行かを判断）
        label = ws.cell(row=row_idx, column=1).value
        if label is not None:
            label_str = str(label).strip()
            # 集計行に達したら終了
            if label_str in ("出席情報",):
                break

        for class_num in CLASS_NUMBERS:
            col = get_class_col(class_num)
            cell_val = ws.cell(row=row_idx, column=col).value
            if is_valid_name(cell_val):
                add_record(
                    name=cell_val,
                    class_name=str(class_num),
                    student_number="",
                    role="教師",
                    memo="担任・副担任",
                )
                teacher_rows_found += 1

    print(f"  教師: {teacher_rows_found} 名")

    # ──────────────────────────────────────────
    # 学年主任（このExcelには存在しないが、
    # 手動追加や別シートからの取込に対応する設計として残す）
    # ──────────────────────────────────────────
    # 学年主任は class_name="" role="学年主任" でCSVに記入してください

    return records


def write_csv(records: list[dict], output_path: str) -> None:
    """CSVファイルに書き出す（BOM付きUTF-8 → Excelで開いても文字化けしない）"""
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["氏名", "メールアドレス", "クラス", "出席番号", "役割", "幹事メモ"])
        for rec in records:
            writer.writerow([
                rec["name"],
                rec["email"],
                rec["class_name"],
                rec["student_number"],
                rec["role"],
                rec["teacher_memo"],
            ])
    print(f"\nCSV出力完了: {output_path} ({len(records)} 件)")


def main():
    if len(sys.argv) < 2:
        print("使い方: python excel_to_roster_csv.py <Excelファイル> [出力CSV]")
        print("例:     python excel_to_roster_csv.py 出席表.xlsx data/roster.csv")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not Path(excel_path).exists():
        print(f"エラー: ファイルが見つかりません: {excel_path}")
        sys.exit(1)

    # 出力先（省略時は同じフォルダに roster.csv）
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        output_path = str(Path(excel_path).parent / "roster.csv")

    records = parse_excel(excel_path)
    write_csv(records, output_path)

    # サマリー表示
    students = sum(1 for r in records if r["role"] == "生徒")
    teachers = sum(1 for r in records if r["role"] == "教師")
    print(f"\n内訳: 生徒 {students} 名 / 教師 {teachers} 名")
    print(f"\n次のステップ:")
    print(f"  1. {output_path} をExcelで開いてメールアドレスを入力してください")
    print(f"  2. 管理画面の「名簿管理」→「CSVを取込む」からアップロードしてください")


if __name__ == "__main__":
    main()
